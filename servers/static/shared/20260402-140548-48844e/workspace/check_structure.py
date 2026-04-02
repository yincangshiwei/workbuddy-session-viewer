from openpyxl import load_workbook
import sys

def check_sheet(sheet_name):
    print(f"\n=== Sheet: {sheet_name} ===")
    wb = load_workbook('Lenta 2025圣诞自留版.xlsx')
    sheet = wb[sheet_name]
    
    print(f"最大行数: {sheet.max_row}")
    print(f"最大列数: {sheet.max_column}")
    
    # 检查前3行数据
    print("\n前3行的前40列:")
    for row_idx in range(1, min(4, sheet.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(41, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is not None:
                row_data.append(f"({col_idx}:{value})")
        print(f"  行{row_idx}: {row_data[:10]}")
    
    wb.close()

# 检查所有需要的sheet
for sheet_name in ["塑料球", "外厂", "陶瓷"]:
    check_sheet(sheet_name)
