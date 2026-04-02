from openpyxl import load_workbook
import sys

sys.stdout.reconfigure(encoding='utf-8')

def check_sheet_data(sheet_name):
    print(f"\n=== Sheet: {sheet_name} ===")
    wb = load_workbook('Lenta 2025圣诞自留版.xlsx')
    sheet = wb[sheet_name]
    
    print(f"第3-20行数据(只显示前10列):")
    for row_idx in range(3, min(21, sheet.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(11, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            row_data.append(str(value)[:20] if value else "")
        # 只显示非空行
        if any(row_data):
            print(f"  行{row_idx}: {row_data}")
    
    wb.close()

# 检查所有需要的sheet
for sheet_name in ["陶瓷"]:
    check_sheet_data(sheet_name)
