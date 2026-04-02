from openpyxl import load_workbook
import sys

sys.stdout.reconfigure(encoding='utf-8')

def check_sheet_data(sheet_name):
    print(f"\n=== Sheet: {sheet_name} ===")
    wb = load_workbook('Lenta 2025圣诞自留版.xlsx')
    sheet = wb[sheet_name]
    
    print(f"第2行表头(列30-40):")
    for col_idx in range(30, min(41, sheet.max_column + 1)):
        cell = sheet.cell(row=2, column=col_idx)
        value = cell.value
        print(f"  列{col_idx}: {value}")
    
    print(f"\n第3-5行数据(列30-40):")
    for row_idx in range(3, min(6, sheet.max_row + 1)):
        row_data = []
        for col_idx in range(30, min(41, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            row_data.append(str(value)[:20] if value else "")
        print(f"  行{row_idx}: {row_data}")
    
    wb.close()

# 检查外厂sheet
check_sheet_data("外厂")
