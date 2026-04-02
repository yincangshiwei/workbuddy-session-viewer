from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import os
import io

def process_excel(source_file, template_file, sheet_name):
    """处理一个sheet的数据填充"""
    print(f"正在处理 sheet: {sheet_name}")
    
    # 加载自留版文件
    source_wb = load_workbook(source_file, data_only=True)
    source_sheet = source_wb[sheet_name]
    
    # 为每个sheet创建一个新的模板文件
    template_wb = load_workbook(template_file)
    
    # 处理第一个sheet (TOTAL PREORDER),从第12行开始
    sheet1 = template_wb["TOTAL PREORDER"]
    start_row = 12
    
    # 获取自留版数据(从第3行开始,第2行是表头)
    source_rows = list(source_sheet.iter_rows(min_row=3, values_only=False))
    
    # 如果没有数据,直接保存并返回
    if len(source_rows) == 0:
        output_file = f"{sheet_name}_填充结果.xlsx"
        template_wb.save(output_file)
        print(f"  Sheet {sheet_name} 没有数据,已保存空文件: {output_file}")
        return output_file
    
    # 遍历每一行数据
    for row_idx, row in enumerate(source_rows):
        target_row = start_row + row_idx
        
        # 如果不是第一行,需要复制上一行的格式
        if row_idx > 0:
            sheet1.insert_rows(target_row)
            # 复制上一行的格式
            prev_row = target_row - 1
            for col_idx, cell in enumerate(sheet1[prev_row], 1):
                col_letter = get_column_letter(col_idx)
                new_cell = sheet1[f"{col_letter}{target_row}"]
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()
        
        # 获取列数据(使用字母索引)
        # B列(2) -> N列(14)
        # D列(4) -> P列(16)
        # F列(6) -> U列(21)
        # N列(14) -> AV列(48)
        # AD列(30) -> BV列(74), 乘以10
        # A列(1) -> BW列(75), 乘以10
        # AF列(32) -> BX列(76), 乘以10
        # AL列(38) -> BY列(77), 乘以1000
        # AK列(37) -> BZ列(78), 乘以1000
        
        B = row[1]
        D = row[3]
        F = row[5]
        N = row[13]
        AD = row[29]
        A = row[0]
        AF = row[31]
        AL = row[37]
        AK = row[36]
        
        # 处理图片(B列)
        if B and hasattr(B, 'value'):
            sheet1[f"N{target_row}"].value = B.value
        elif B:
            sheet1[f"N{target_row}"].value = B
        
        # 填充其他列
        if D is not None:
            sheet1[f"P{target_row}"].value = D
        if F is not None:
            sheet1[f"U{target_row}"].value = F
        if N is not None:
            sheet1[f"AV{target_row}"].value = N
        
        # 需要乘以10的列
        if AD is not None:
            try:
                sheet1[f"BV{target_row}"].value = float(AD) * 10
            except:
                sheet1[f"BV{target_row}"].value = AD
        if A is not None:
            try:
                sheet1[f"BW{target_row}"].value = float(A) * 10
            except:
                sheet1[f"BW{target_row}"].value = A
        if AF is not None:
            try:
                sheet1[f"BX{target_row}"].value = float(AF) * 10
            except:
                sheet1[f"BX{target_row}"].value = AF
        
        # 需要乘以1000的列
        if AL is not None:
            try:
                sheet1[f"BY{target_row}"].value = float(AL) * 1000
            except:
                sheet1[f"BY{target_row}"].value = AL
        if AK is not None:
            try:
                sheet1[f"BZ{target_row}"].value = float(AK) * 1000
            except:
                sheet1[f"BZ{target_row}"].value = AK
        
        # 提取并插入图片(B列的图片)
        for image in source_sheet._images:
            if image.anchor._from:
                img_row = image.anchor._from.row
                if img_row == 2 + row_idx:  # 第3行开始(索引2)
                    # 添加图片到目标单元格
                    new_img = Image(image.ref)
                    new_img.anchor = sheet1.cell(row=target_row, column=14)  # N列
                    sheet1.add_image(new_img)
                    break
    
    # 处理第二个sheet (DC),从第10行开始
    sheet2 = template_wb["DC"]
    start_row_2 = 10
    
    for row_idx, row in enumerate(source_rows):
        target_row = start_row_2 + row_idx
        
        # 如果不是第一行,需要复制上一行的格式
        if row_idx > 0:
            sheet2.insert_rows(target_row)
            # 复制上一行的格式
            prev_row = target_row - 1
            for col_idx, cell in enumerate(sheet2[prev_row], 1):
                col_letter = get_column_letter(col_idx)
                new_cell = sheet2[f"{col_letter}{target_row}"]
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()
        
        # N列(14) -> D列(4)
        N_value = row[13]
        if N_value is not None:
            sheet2[f"D{target_row}"].value = N_value
    
    # 保存文件
    output_file = f"{sheet_name}_填充结果.xlsx"
    template_wb.save(output_file)
    print(f"  已保存: {output_file}")
    return output_file

def main():
    source_file = "Lenta 2025圣诞自留版.xlsx"
    template_file = "（Lenta）Preorder. Эталонная форма.xlsx"
    
    # 需要处理的sheet名称
    sheets_to_process = ["塑料球", "外厂", "陶瓷"]
    
    output_files = []
    for sheet_name in sheets_to_process:
        try:
            output_file = process_excel(source_file, template_file, sheet_name)
            output_files.append(output_file)
        except Exception as e:
            print(f"处理 {sheet_name} 时出错: {str(e)}")
            import traceback
            traceback.print_exc()
    
    print("\n处理完成!生成的文件:")
    for f in output_files:
        print(f"  - {f}")

if __name__ == "__main__":
    main()
