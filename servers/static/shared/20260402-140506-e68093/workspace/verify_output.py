from openpyxl import load_workbook
import sys

sys.stdout.reconfigure(encoding='utf-8')

def verify_file(filename):
    print(f"\n=== 验证文件: {filename} ===")
    wb = load_workbook(filename)
    
    # 检查第一个sheet (TOTAL PREORDER)
    sheet1 = wb["TOTAL PREORDER"]
    print(f"\nSheet1 (TOTAL PREORDER):")
    print(f"  总行数: {sheet1.max_row}")
    
    # 检查第12行开始的几行数据
    print(f"\n  前3行数据(从第12行开始):")
    for row_idx in range(12, min(15, sheet1.max_row + 1)):
        # 检查关键列
        n_cell = sheet1[f"N{row_idx}"]
        p_cell = sheet1[f"P{row_idx}"]
        u_cell = sheet1[f"U{row_idx}"]
        av_cell = sheet1[f"AV{row_idx}"]
        bv_cell = sheet1[f"BV{row_idx}"]
        bw_cell = sheet1[f"BW{row_idx}"]
        bx_cell = sheet1[f"BX{row_idx}"]
        by_cell = sheet1[f"BY{row_idx}"]
        bz_cell = sheet1[f"BZ{row_idx}"]
        
        print(f"    行{row_idx}: N={n_cell.value}, P={p_cell.value}, U={u_cell.value}, AV={av_cell.value}, BV={bv_cell.value}, BW={bw_cell.value}, BX={bx_cell.value}, BY={by_cell.value}, BZ={bz_cell.value}")
    
    # 检查第二个sheet (DC)
    sheet2 = wb["DC"]
    print(f"\nSheet2 (DC):")
    print(f"  总行数: {sheet2.max_row}")
    
    # 检查第10行开始的几行数据
    print(f"\n  前3行数据(从第10行开始):")
    for row_idx in range(10, min(13, sheet2.max_row + 1)):
        d_cell = sheet2[f"D{row_idx}"]
        print(f"    行{row_idx}: D={d_cell.value}")
    
    # 检查图片
    print(f"\nSheet1 图片数量: {len(sheet1._images)}")
    
    wb.close()

# 验证所有生成的文件
files = ["塑料球_填充结果.xlsx", "外厂_填充结果.xlsx", "陶瓷_填充结果.xlsx"]
for f in files:
    try:
        verify_file(f)
    except Exception as e:
        print(f"验证 {f} 时出错: {e}")

print("\n" + "="*50)
print("验证完成!")
print("="*50)
