from openpyxl import load_workbook
import sys

sys.stdout.reconfigure(encoding='utf-8')

def check_columns(sheet_name):
    print(f"\n=== Sheet: {sheet_name} ===")
    wb = load_workbook('Lenta 2025圣诞自留版.xlsx')
    sheet = wb[sheet_name]
    
    print(f"最大列数: {sheet.max_column}")
    
    # 需要的列索引(从1开始)
    required_cols = {
        1: 'A',
        2: 'B',
        4: 'D',
        6: 'F',
        14: 'N',
        30: 'AD',
        32: 'AF',
        37: 'AK',
        38: 'AL'
    }
    
    print("\n需要的列是否存在:")
    for col_idx, col_name in sorted(required_cols.items()):
        if col_idx <= sheet.max_column:
            cell = sheet.cell(row=3, column=col_idx)
            print(f"  {col_name}列(索引{col_idx}): 存在, 示例值={cell.value}")
        else:
            print(f"  {col_name}列(索引{col_idx}): 不存在!")
    
    wb.close()

# 检查所有需要的sheet
for sheet_name in ["塑料球", "外厂", "陶瓷"]:
    check_columns(sheet_name)
