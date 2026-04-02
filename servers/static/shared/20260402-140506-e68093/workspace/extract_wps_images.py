# -*- coding: utf-8 -*-
"""
提取WPS DISPIMG图片并填充到模板Excel
"""
import zipfile
import xml.etree.ElementTree as ET
import re
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D, XDRPoint2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
import os
import shutil
from io import BytesIO

def extract_wps_dispimg_images(xlsx_file):
    """
    从WPS Excel文件中提取DISPIMG图片
    
    Args:
        xlsx_file: Excel文件路径
    
    Returns:
        dict: {DISPIMG_ID: (image_data, extension)}
    """
    images = {}
    
    try:
        # Excel文件本质是ZIP压缩包
        with zipfile.ZipFile(xlsx_file, 'r') as zip_file:
            # 1. 解析 cellimages.xml
            cell_images_xml = None
            rels_xml = None
            
            try:
                cell_images_xml = zip_file.read('xl/cellimages.xml')
            except KeyError:
                print(f"未找到 xl/cellimages.xml 文件")
                return images
            
            try:
                rels_xml = zip_file.read('xl/_rels/cellimages.xml.rels')
            except KeyError:
                print(f"未找到 xl/_rels/cellimages.xml.rels 文件")
                return images
            
            # 2. 解析cellimages.xml，建立ID与rId的映射
            cell_images_root = ET.fromstring(cell_images_xml)
            
            # 命名空间
            namespaces = {
                'etc': 'http://schemas.microsoft.com/office/office/2006/relationships',
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            }
            
            # 解析rels文件，建立rId与图片路径的映射
            rels_root = ET.fromstring(rels_xml)
            r_id_to_path = {}
            
            for rel in rels_root.findall('.//{*}Relationship'):
                r_id = rel.get('Id')
                target = rel.get('Target')
                if r_id and target:
                    r_id_to_path[r_id] = target
            
            # 3. 提取图片
            for pic in cell_images_root.findall('.//xdr:pic', namespaces):
                try:
                    # 获取DISPIMG ID
                    cNvPr = pic.find('.//xdr:cNvPr', namespaces)
                    if cNvPr is None:
                        continue
                    
                    dispimg_id = cNvPr.get('name')
                    if not dispimg_id or not dispimg_id.startswith('ID_'):
                        continue
                    
                    # 获取rId
                    blip = pic.find('.//a:blip', namespaces)
                    if blip is None:
                        continue
                    
                    r_id = blip.get(f'{{{namespaces["r"]}}}embed')
                    if not r_id or r_id not in r_id_to_path:
                        continue
                    
                    # 获取图片在ZIP中的路径
                    img_path = 'xl/' + r_id_to_path[r_id].lstrip('/')
                    
                    # 读取图片数据
                    img_data = zip_file.read(img_path)
                    
                    # 确定图片扩展名
                    ext = 'png'
                    if 'jpeg' in img_path.lower() or 'jpg' in img_path.lower():
                        ext = 'jpeg'
                    elif 'png' in img_path.lower():
                        ext = 'png'
                    elif 'gif' in img_path.lower():
                        ext = 'gif'
                    
                    images[dispimg_id] = (img_data, ext)
                    
                except Exception as e:
                    print(f"提取图片时出错: {e}")
                    continue
    
    except Exception as e:
        print(f"处理ZIP文件时出错: {e}")
    
    return images


def get_dispimg_ids_from_cells(sheet, start_row=3):
    """
    从Excel sheet中提取DISPIMG函数中的ID
    
    Args:
        sheet: openpyxl工作表对象
        start_row: 数据开始行号
    
    Returns:
        dict: {(row_idx, col_idx): DISPIMG_ID}
    """
    dispimg_map = {}
    
    for row in sheet.iter_rows(min_row=start_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # 匹配 =_xlfn.DISPIMG("ID_xxxx",1) 或 =DISPIMG("ID_xxxx",1)
                match = re.search(r'DISPIMG\s*\(\s*"([^"]+)"', cell.value)
                if match:
                    dispimg_id = match.group(1)
                    dispimg_map[(cell.row, cell.column)] = dispimg_id
    
    return dispimg_map


def fill_template_with_images(source_file, template_file, output_file, sheet_name):
    """
    填充模板Excel文件，包括图片
    
    Args:
        source_file: 源文件路径
        template_file: 模板文件路径
        output_file: 输出文件路径
        sheet_name: 要处理的sheet名称
    """
    print(f"\n处理sheet: {sheet_name}")
    
    # 1. 从源文件提取所有DISPIMG图片
    print(f"从源文件提取图片: {source_file}")
    source_wb = openpyxl.load_workbook(source_file, data_only=False)
    source_sheet = source_wb[sheet_name]
    
    # 获取DISPIMG ID映射
    dispimg_map = get_dispimg_ids_from_cells(source_sheet, start_row=3)
    print(f"找到 {len(dispimg_map)} 个DISPIMG单元格")
    
    # 提取图片数据
    images = extract_wps_dispimg_images(source_file)
    print(f"提取到 {len(images)} 个图片")
    
    # 2. 打开模板文件
    print(f"打开模板文件: {template_file}")
    template_wb = openpyxl.load_workbook(template_file)
    
    # Sheet1: TOTAL PREORDER (从第12行开始)
    sheet1 = template_wb['TOTAL PREORDER']
    # Sheet2: DC (从第10行开始)
    sheet2 = template_wb['DC']
    
    # 3. 读取源数据
    source_data = []
    for row in source_sheet.iter_rows(min_row=3):
        # 检查是否为空行（前几列都为空）
        is_empty = True
        for cell in row[:5]:  # 检查前5列
            if cell.value is not None and cell.value != '':
                is_empty = False
                break
        if not is_empty:
            source_data.append(row)
    
    print(f"源数据行数: {len(source_data)}")
    
    # 4. 填充Sheet1 (TOTAL PREORDER) - 从第12行开始
    start_row_sheet1 = 12
    col_mapping = {
        'B': 'N',   # 图片
        'D': 'P',   # 客户货号
        'F': 'U',   # 产品描述
        'N': 'AV',  # N列
        'AD': 'BV', # ×10
        'A': 'BW',  # ×10
        'AF': 'BX', # ×10
        'AK': 'BZ', # ×1000
        'AL': 'BY', # ×1000
    }
    
    # 先复制列宽
    for col_idx in range(1, sheet1.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if col_letter in sheet1.column_dimensions:
            sheet1.column_dimensions[col_letter].width = sheet1.column_dimensions[col_letter].width

    # 获取模板行的行高
    template_row_height = sheet1.row_dimensions[start_row_sheet1].height

    for i, source_row in enumerate(source_data):
        target_row = start_row_sheet1 + i

        # 设置行高
        if template_row_height:
            sheet1.row_dimensions[target_row].height = template_row_height

        # 复制格式
        for col in range(1, sheet1.max_column + 1):
            source_cell = sheet1.cell(row=start_row_sheet1, column=col)
            target_cell = sheet1.cell(row=target_row, column=col)

            # 复制样式
            if source_cell.has_style:
                target_cell.font = source_cell.font.copy()
                target_cell.border = source_cell.border.copy()
                target_cell.fill = source_cell.fill.copy()
                target_cell.number_format = source_cell.number_format
                target_cell.protection = source_cell.protection.copy()
                target_cell.alignment = source_cell.alignment.copy()

        # 填充数据
        for src_col_letter, dst_col_letter in col_mapping.items():
            src_col_idx = openpyxl.utils.column_index_from_string(src_col_letter) - 1  # 0-based
            dst_col_idx = openpyxl.utils.column_index_from_string(dst_col_letter)       # 1-based
            
            if src_col_idx < len(source_row):
                src_cell = source_row[src_col_idx]
                dst_cell = sheet1.cell(row=target_row, column=dst_col_idx)
                
                # 特殊处理：B列图片
                if src_col_letter == 'B':
                    # 检查是否有DISPIMG公式
                    if src_cell.value and isinstance(src_cell.value, str):
                        match = re.search(r'DISPIMG\s*\(\s*"([^"]+)"', src_cell.value)
                        if match:
                            dispimg_id = match.group(1)
                            if dispimg_id in images:
                                # 使用BytesIO避免临时文件
                                img_data, ext = images[dispimg_id]
                                img_stream = BytesIO(img_data)

                                try:
                                    # 插入图片到单元格
                                    img = Image(img_stream)

                                    # 获取单元格的尺寸
                                    # 列宽转换为像素: 1列宽单位 ≈ 7像素
                                    col_width = sheet1.column_dimensions[dst_col_letter].width
                                    if col_width is None:
                                        col_width = 8.43  # Excel默认列宽
                                    col_width_pixels = col_width * 7

                                    # 行高转换为像素: 1行高单位 ≈ 1.35像素
                                    row_height = sheet1.row_dimensions[target_row].height or 15
                                    row_height_pixels = row_height * 1.35

                                    # 获取原始图片尺寸
                                    original_width = img.width
                                    original_height = img.height
                                    original_ratio = original_width / original_height

                                    # 计算缩放比例，使图片适应单元格
                                    # 保持宽高比,取较小的缩放比例
                                    width_ratio = col_width_pixels / original_width
                                    height_ratio = row_height_pixels / original_height
                                    scale = min(width_ratio, height_ratio) * 0.9  # 保留10%边距

                                    # 设置缩放后的尺寸
                                    img.width = int(original_width * scale)
                                    img.height = int(original_height * scale)

                                    # 创建自定义锚点,将图片居中到单元格
                                    # 计算单元格在像素中的位置
                                    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
                                    from openpyxl.utils.units import pixels_to_EMU

                                    # 计算居中偏移
                                    offset_x_pixels = (col_width_pixels - img.width) / 2
                                    offset_y_pixels = (row_height_pixels - img.height) / 2

                                    # 设置锚点到单元格
                                    img.anchor = dst_cell.coordinate

                                    sheet1.add_image(img)

                                    print(f"  [OK] 第{target_row}行: 插入图片 {dispimg_id} (尺寸: {img.width}x{img.height}, 单元格: {col_width_pixels:.0f}x{row_height_pixels:.0f})")
                                except Exception as e:
                                    print(f"  [FAIL] 第{target_row}行: 插入图片失败 {e}")
                elif src_col_letter in ['AD', 'A', 'AF']:
                    # ×10
                    if src_cell.value is not None:
                        try:
                            val = float(src_cell.value)
                            dst_cell.value = val * 10
                        except (ValueError, TypeError):
                            dst_cell.value = src_cell.value
                elif src_col_letter in ['AK', 'AL']:
                    # ×1000
                    if src_cell.value is not None:
                        try:
                            val = float(src_cell.value)
                            dst_cell.value = val * 1000
                        except (ValueError, TypeError):
                            dst_cell.value = src_cell.value
                else:
                    # 普通复制
                    dst_cell.value = src_cell.value
    
    # 5. 填充Sheet2 (DC) - 从第10行开始
    start_row_sheet2 = 10

    # 复制Sheet2的列宽
    for col_idx in range(1, sheet2.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if col_letter in sheet2.column_dimensions:
            sheet2.column_dimensions[col_letter].width = sheet2.column_dimensions[col_letter].width

    # 获取模板行的行高
    template_row_height_sheet2 = sheet2.row_dimensions[start_row_sheet2].height

    for i, source_row in enumerate(source_data):
        target_row = start_row_sheet2 + i

        # 设置行高
        if template_row_height_sheet2:
            sheet2.row_dimensions[target_row].height = template_row_height_sheet2

        # 复制格式
        for col in range(1, sheet2.max_column + 1):
            source_cell = sheet2.cell(row=start_row_sheet2, column=col)
            target_cell = sheet2.cell(row=target_row, column=col)

            # 复制样式
            if source_cell.has_style:
                target_cell.font = source_cell.font.copy()
                target_cell.border = source_cell.border.copy()
                target_cell.fill = source_cell.fill.copy()
                target_cell.number_format = source_cell.number_format
                target_cell.protection = source_cell.protection.copy()
                target_cell.alignment = source_cell.alignment.copy()
        
        # N列 -> D列
        src_col_idx = 13  # N列 (0-based)
        dst_col_idx = 4   # D列 (1-based)
        
        if src_col_idx < len(source_row):
            src_cell = source_row[src_col_idx]
            dst_cell = sheet2.cell(row=target_row, column=dst_col_idx)
            
            # 如果是DISPIMG公式，清空（因为DC sheet不需要图片）
            if src_cell.value and isinstance(src_cell.value, str):
                if 'DISPIMG' in src_cell.value:
                    dst_cell.value = ''
                else:
                    dst_cell.value = src_cell.value
            else:
                dst_cell.value = src_cell.value
    
    # 6. 保存文件
    print(f"保存文件: {output_file}")
    template_wb.save(output_file)
    print(f"[OK] 完成!")
    
    source_wb.close()
    template_wb.close()


def main():
    # 文件路径
    source_file = "Lenta 2025圣诞自留版.xlsx"
    template_file = "（Lenta）Preorder. Эталонная форма.xlsx"
    
    # 要处理的sheet
    sheets = ['塑料球', '外厂', '陶瓷']
    
    for sheet_name in sheets:
        output_file = f"{sheet_name}_填充结果_含图片.xlsx"
        print(f"\n{'='*60}")
        print(f"开始处理: {sheet_name}")
        print(f"{'='*60}")
        
        fill_template_with_images(source_file, template_file, output_file, sheet_name)


if __name__ == '__main__':
    main()
