from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.utils import get_column_letter
import sys
import copy

sys.stdout.reconfigure(encoding='utf-8')

def get_cell_value(row, col_idx):
    """安全地获取单元格值,处理Cell对象"""
    if col_idx >= len(row):
        return None
    cell = row[col_idx]
    if hasattr(cell, 'value'):
        return cell.value
    return cell

def is_image_formula(value):
    """检查是否是WPS图片公式"""
    if value and isinstance(value, str):
        return value.startswith('=_xlfn.DISPIMG(')
    return False

def process_excel(source_file, template_file, sheet_name):
    """处理一个sheet的数据填充"""
    print(f"\n正在处理 sheet: {sheet_name}")
    
    # 加载自留版文件
    source_wb = load_workbook(source_file, data_only=True)
    source_sheet = source_wb[sheet_name]
    
    # 为每个sheet创建一个新的模板文件
    template_wb = load_workbook(template_file)
    
    # 处理第一个sheet (TOTAL PREORDER),从第12行开始
    sheet1 = template_wb["TOTAL PREORDER"]
    start_row = 12
    
    # 获取自留版数据(从第3行开始,第2行是表头)
    source_rows = list(source_sheet.iter_rows(min_row=3, values_only=False))
    
    # 如果没有数据,直接保存并返回
    if len(source_rows) == 0:
        output_file = f"{sheet_name}_填充结果.xlsx"
        template_wb.save(output_file)
        print(f"  Sheet {sheet_name} 没有数据,已保存空文件: {output_file}")
        source_wb.close()
        template_wb.close()
        return output_file
    
    print(f"  共有 {len(source_rows)} 行数据")
    
    # 收集每行的图片
    images_by_row = {}
    for img in source_sheet._images:
        if hasattr(img.anchor, '_from') and img.anchor._from:
            img_row = img.anchor._from.row + 1  # 转换为1-based索引
            if img_row >= 3:  # 从第3行开始
                images_by_row[img_row] = img
    
    # 遍历每一行数据
    for row_idx, row in enumerate(source_rows):
        target_row = start_row + row_idx
        source_row_num = row_idx + 3  # 源文件的实际行号(从3开始)
        
        # 如果不是第一行,需要复制上一行的格式
        if row_idx > 0:
            sheet1.insert_rows(target_row)
            # 复制上一行的格式
            prev_row = target_row - 1
            for col_idx in range(1, sheet1.max_column + 1):
                cell = sheet1.cell(row=prev_row, column=col_idx)
                new_cell = sheet1.cell(row=target_row, column=col_idx)
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)
        
        # 获取列数据(列索引从0开始)
        B_val = get_cell_value(row, 1)
        D_val = get_cell_value(row, 3)
        F_val = get_cell_value(row, 5)
        N_val = get_cell_value(row, 13)
        AD_val = get_cell_value(row, 29)
        A_val = get_cell_value(row, 0)
        AF_val = get_cell_value(row, 31)
        AL_val = get_cell_value(row, 37)
        AK_val = get_cell_value(row, 36)
        
        # 处理B列(图片或文本)
        if B_val is not None:
            if is_image_formula(B_val):
                # 如果是图片公式,不填充值(需要从图片对象中提取)
                sheet1[f"N{target_row}"].value = ""
            else:
                sheet1[f"N{target_row}"].value = B_val
        
        # 填充其他列
        if D_val is not None:
            sheet1[f"P{target_row}"].value = D_val
        if F_val is not None:
            sheet1[f"U{target_row}"].value = F_val
        if N_val is not None:
            sheet1[f"AV{target_row}"].value = N_val
        
        # 需要乘以10的列
        if AD_val is not None:
            try:
                sheet1[f"BV{target_row}"].value = float(AD_val) * 10
            except (ValueError, TypeError):
                sheet1[f"BV{target_row}"].value = AD_val
        
        if A_val is not None:
            try:
                sheet1[f"BW{target_row}"].value = float(A_val) * 10
            except (ValueError, TypeError):
                sheet1[f"BW{target_row}"].value = A_val
        
        if AF_val is not None:
            try:
                sheet1[f"BX{target_row}"].value = float(AF_val) * 10
            except (ValueError, TypeError):
                sheet1[f"BX{target_row}"].value = AF_val
        
        # 需要乘以1000的列
        if AL_val is not None:
            try:
                sheet1[f"BY{target_row}"].value = float(AL_val) * 1000
            except (ValueError, TypeError):
                sheet1[f"BY{target_row}"].value = AL_val
        
        if AK_val is not None:
            try:
                sheet1[f"BZ{target_row}"].value = float(AK_val) * 1000
            except (ValueError, TypeError):
                sheet1[f"BZ{target_row}"].value = AK_val
        
        # 添加图片(如果有)
        if source_row_num in images_by_row:
            try:
                source_img = images_by_row[source_row_num]
                new_img = Image(source_img.ref)
                # 使用正确的锚点方式
                marker = AnchorMarker(col=13, colOff=0, row=target_row-1, rowOff=0)
                new_img.anchor = TwoCellAnchor(_from=marker, to=None)
                sheet1.add_image(new_img)
            except Exception as e:
                print(f"    警告: 第{target_row}行图片添加失败: {e}")
    
    # 处理第二个sheet (DC),从第10行开始
    sheet2 = template_wb["DC"]
    start_row_2 = 10
    
    for row_idx, row in enumerate(source_rows):
        target_row = start_row_2 + row_idx
        
        # 如果不是第一行,需要复制上一行的格式
        if row_idx > 0:
            sheet2.insert_rows(target_row)
            # 复制上一行的格式
            prev_row = target_row - 1
            for col_idx in range(1, sheet2.max_column + 1):
                cell = sheet2.cell(row=prev_row, column=col_idx)
                new_cell = sheet2.cell(row=target_row, column=col_idx)
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)
        
        # N列(索引13) -> D列(4)
        N_value = get_cell_value(row, 13)
        if N_value is not None:
            sheet2[f"D{target_row}"].value = N_value
    
    # 保存文件
    output_file = f"{sheet_name}_填充结果.xlsx"
    template_wb.save(output_file)
    print(f"  已保存: {output_file}")
    
    source_wb.close()
    template_wb.close()
    return output_file

def main():
    source_file = "Lenta 2025圣诞自留版.xlsx"
    template_file = "（Lenta）Preorder. Эталонная форма.xlsx"
    
    # 需要处理的sheet名称
    sheets_to_process = ["塑料球", "外厂", "陶瓷"]
    
    output_files = []
    for sheet_name in sheets_to_process:
        try:
            output_file = process_excel(source_file, template_file, sheet_name)
            output_files.append(output_file)
        except Exception as e:
            print(f"\n处理 {sheet_name} 时出错: {str(e)}")
            import traceback
            traceback.print_exc()
    
    print("\n" + "="*50)
    print("处理完成!生成的文件:")
    for f in output_files:
        print(f"  - {f}")
    print("="*50)

if __name__ == "__main__":
    main()
